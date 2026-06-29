"""Generate a sample bulk-import bundle (.zip) for testing Settings > Import.

Produces ``khata-sample-import.zip`` containing a filled workbook plus an
``images/`` folder of labelled placeholder photos. The workbook mirrors the
real template (see ``app.services.import_template.SHEETS``) and uses the
default seeded dropdown values so it validates and commits cleanly on a fresh DB.

Run:  .venv/bin/python scripts/make_sample_import.py [output.zip]
"""
from __future__ import annotations

import io
import sys
import zipfile
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw

from app.services.import_template import INSTRUCTIONS, SHEETS

HEADER_FILL = PatternFill("solid", fgColor="1C1B19")
HEADER_FONT = Font(color="FBF8F2", bold=True)

TODAY = date.today()


def d(days_ago: int) -> str:
    return (TODAY - timedelta(days=days_ago)).isoformat()


# ----- Sample rows (keyed by sheet; values follow SHEETS column order) --------
# Dropdown cells use default-seeded names (categories, purities, weight types,
# supply sources, order sources) so a fresh DB accepts them as-is.
SAMPLE = {
    "Customers": [
        {"name": "Aarti Sharma", "phone": "98100 11223", "address": "Vasundhara, Ghaziabad", "notes": "Prefers 22 KT"},
        {"name": "Rohit Mehta", "phone": "98200 44556", "address": "Indirapuram", "notes": ""},
        {"name": "Priya Nair", "phone": "", "address": "", "notes": "Referred by Aarti"},
    ],
    "Parties": [
        {"name": "Verma Bullion", "phone": "98300 77889", "address": "Karol Bagh, Delhi", "notes": "Gold supplier"},
        {"name": "Shree Diamonds", "phone": "", "address": "Surat", "notes": "Loose stones"},
    ],
    "Opening Balances": [
        {"entity_type": "customer", "entity_name": "Aarti Sharma", "as_of_date": d(120),
         "amount": "15000", "direction": "debit"},
        {"entity_type": "party", "entity_name": "Verma Bullion", "as_of_date": d(120),
         "amount": "40000", "direction": "credit"},
    ],
    "Orders": [
        # Plain gold ring — metal only. images: two photos of the same ring.
        {"order_ref": "S-001", "customer_name": "Aarti Sharma", "order_date": d(20),
         "item_category": "Ring", "item_name": "Engagement band", "weight_type": "Lightweight",
         "supply_source": "On Order", "purity": "22 KT", "source": "Whatsapp",
         "reference": "Friend", "gross_weight": "6.500", "diamond_weight": "", "diamond_rate": "",
         "stone_weight": "", "stone_rate": "", "others_weight": "", "others_rate": "",
         "metal_rate": "6200", "labour_rate": "600", "order_code": "RNG-22-001",
         "status": "delivered", "payment_received": "30000", "payment_mode": "upi",
         "notes": "Size 12", "images": "ring1.jpg;ring1-side.jpg"},
        # Diamond necklace — gross + diamond carats. one photo.
        {"order_ref": "S-002", "customer_name": "Rohit Mehta", "order_date": d(12),
         "item_category": "Necklace", "item_name": "Solitaire necklace", "weight_type": "Heavyweight",
         "supply_source": "On Order", "purity": "18 KT", "source": "Instagram",
         "reference": "", "gross_weight": "18.000", "diamond_weight": "2.50", "diamond_rate": "55000",
         "stone_weight": "", "stone_rate": "", "others_weight": "", "others_rate": "",
         "metal_rate": "5800", "labour_rate": "1200", "order_code": "NCK-18-014",
         "status": "pending", "payment_received": "50000", "payment_mode": "bank_transfer",
         "notes": "Deliver before Diwali", "images": "necklace1.png"},
        # Silver bangle pair — no photos (tests optional images).
        {"order_ref": "S-003", "customer_name": "Priya Nair", "order_date": d(5),
         "item_category": "Bangle", "item_name": "Oxidised bangles", "weight_type": "Normal",
         "supply_source": "Stock", "purity": "Silver", "source": "Walk-in",
         "reference": "", "gross_weight": "45.000", "diamond_weight": "", "diamond_rate": "",
         "stone_weight": "3.00", "stone_rate": "400", "others_weight": "", "others_rate": "",
         "metal_rate": "95", "labour_rate": "40", "order_code": "",
         "status": "delivered", "payment_received": "0", "payment_mode": "",
         "notes": "", "images": ""},
    ],
    "Cash Entries": [
        {"date": d(10), "person_name": "Aarti Sharma", "details": "Balance on ring", "type": "received", "amount": "8500"},
        {"date": d(7), "person_name": "Office", "details": "Misc expenses", "type": "paid", "amount": "1200"},
    ],
    "Purchases": [
        {"date": d(15), "party_name": "Verma Bullion", "details": "100g 22KT bar", "entry_notes": "Rate 6150",
         "amount": "615000", "amount_paid": "300000"},
        {"date": d(8), "party_name": "Shree Diamonds", "details": "Loose solitaires", "entry_notes": "",
         "amount": "180000", "amount_paid": "180000"},
    ],
}

# Photos to generate: filename -> (label, RGB background).
IMAGES = {
    "ring1.jpg": ("Ring — front", (233, 220, 204)),
    "ring1-side.jpg": ("Ring — side", (227, 236, 232)),
    "necklace1.png": ("Necklace", (243, 226, 224)),
}


def make_photo(label: str, bg: tuple[int, int, int], fmt: str) -> bytes:
    img = Image.new("RGB", (640, 480), bg)
    draw = ImageDraw.Draw(img)
    draw.rectangle([20, 20, 619, 459], outline=(28, 27, 25), width=4)
    draw.text((40, 220), label, fill=(28, 27, 25))
    draw.text((40, 250), "Khata sample image", fill=(107, 100, 89))
    buf = io.BytesIO()
    img.save(buf, format=fmt)
    return buf.getvalue()


def build_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    for i, line in enumerate(INSTRUCTIONS, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100

    for title, headers in SHEETS.items():
        sheet = wb.create_sheet(title)
        for col, name in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            sheet.column_dimensions[cell.column_letter].width = max(14, len(name) + 2)
        for r, row in enumerate(SAMPLE.get(title, []), start=2):
            for col, name in enumerate(headers, start=1):
                sheet.cell(row=r, column=col, value=row.get(name, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> None:
    out = sys.argv[1] if len(sys.argv) > 1 else "khata-sample-import.zip"
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("khata-sample-import.xlsx", build_workbook())
        for fname, (label, bg) in IMAGES.items():
            fmt = "PNG" if fname.lower().endswith(".png") else "JPEG"
            zf.writestr(f"images/{fname}", make_photo(label, bg, fmt))
    print(f"Wrote {out}")
    print("  - khata-sample-import.xlsx (Instructions + 6 data sheets)")
    print(f"  - images/ ({len(IMAGES)} photos)")
    print("Upload it under Settings > Import.")


if __name__ == "__main__":
    main()
