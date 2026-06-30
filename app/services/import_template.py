"""Generate the downloadable Excel import template (one sheet per data group)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models import (
    ItemCategory,
    OrderSource,
    PurityType,
    SupplySource,
    WeightType,
)

HEADER_FILL = PatternFill("solid", fgColor="1C1B19")
HEADER_FONT = Font(color="FBF8F2", bold=True)

SHEETS = {
    "Customers": ["name", "phone", "address", "notes"],
    "Parties": ["name", "phone", "address", "notes"],
    "Opening Balances": ["entity_type", "entity_name", "as_of_date", "amount", "direction"],
    "Orders": ["order_ref", "customer_name", "order_date", "item_category", "item_name",
               "weight_type", "supply_source", "purity", "source", "reference",
               "gross_weight", "diamond_weight", "diamond_rate", "stone_weight", "stone_rate",
               "others_weight", "others_rate", "metal_rate", "labour_rate",
               "order_code", "status", "payment_received", "payment_mode", "notes", "images"],
    "Cash Entries": ["date", "person_name", "details", "type", "amount"],
    "Purchases": ["date", "party_name", "details", "entry_notes", "amount", "amount_paid"],
}

INSTRUCTIONS = [
    "Khata — Import Template",
    "",
    "Fill each sheet and upload it under Settings > Import. Validation runs before anything is saved.",
    "",
    "Customers / Parties: 'name' is required; phone/address/notes optional.",
    "Orders: one row per item/piece. item_category is REQUIRED on every row (dropdown).",
    "  Multiple items in one order: give those rows the SAME 'order_ref' (any text, e.g.",
    "  ORD-1001) and keep them on CONSECUTIVE rows (right under each other — rows with the",
    "  same order_ref that are split apart become separate orders). The FIRST row of a",
    "  group carries the order-level fields (customer_name,",
    "  order_date, source, reference, order_code, status, payment_received, payment_mode,",
    "  notes); continuation rows leave those blank and fill only the per-piece columns.",
    "  Leave 'order_ref' blank for a normal single-item order.",
    "  item_name (free text), weight_type, supply_source, purity and 'source' are optional.",
    "  'reference' is free text (e.g. friends/family).",
    "  Item price is computed: gross_weight (g); diamond/stone/others weights are in CARATS",
    "  (5 ct = 1 g). Net (metal) weight = gross − (diamond+stone+others)/5. Price =",
    "  net×metal_rate + diamond_ct×diamond_rate + stone_ct×stone_rate + others_ct×others_rate",
    "  + net×labour_rate. Leave rates blank for parts you don't use.",
    "  diamond_weight/diamond_rate import as one diamond line typed 'Diamond (Other",
    "  fancy)'; re-type it (or add more diamond lines) in the app after importing.",
    "  status = pending or delivered. payment_mode = cash/upi/bank_transfer/old_gold_exchange/other.",
    "  An order's total is the sum of its item subtotals; payment_received applies once per order.",
    "Pictures (optional): list filenames in the 'images' column, separated by ; (e.g.",
    "  ring1.jpg;ring1b.jpg). Put the photos in a folder named 'images', then ZIP this",
    "  filled workbook together with that folder and upload the .zip (instead of the .xlsx).",
    "  Allowed: png/jpg/jpeg/webp/gif, up to 12 per item.",
    "Cash Entries: type = received or paid.",
    "Opening Balances: entity_type = customer or party; direction = debit or credit.",
    "Dates: use YYYY-MM-DD (e.g. 2026-06-14).",
    "Customer/supplier names are matched case-insensitively; new names are created automatically.",
]


def _list_validation(values: list[str]) -> DataValidation:
    joined = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
    return dv


def _style_header(ws, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 2)


def build_template(session: Session) -> bytes:
    purities = [p.name for p in session.query(PurityType).filter_by(is_active=True).all()]
    categories = [c.name for c in session.query(ItemCategory).filter_by(is_active=True).all()]
    weights = [w.name for w in session.query(WeightType).filter_by(is_active=True).all()]
    supplies = [s.name for s in session.query(SupplySource).filter_by(is_active=True).all()]
    sources = [s.name for s in session.query(OrderSource).filter_by(is_active=True).all()]

    wb = Workbook()
    # Instructions sheet first.
    ws_info = wb.active
    ws_info.title = "Instructions"
    for i, line in enumerate(INSTRUCTIONS, start=1):
        ws_info.cell(row=i, column=1, value=line)
    ws_info.column_dimensions["A"].width = 100

    for title, headers in SHEETS.items():
        ws = wb.create_sheet(title)
        _style_header(ws, headers)

        def add_dv(values, col_letter):
            if not values:
                return
            dv = _list_validation(values)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        if title == "Orders":
            # Columns: A ref, B customer, C date, D category, E item_name, F weight_type,
            #          G supply, H purity, I source, J reference, K gross_wt, L dia_wt,
            #          M dia_rate, N stone_wt, O stone_rate, P others_wt, Q others_rate,
            #          R metal_rate, S labour_rate, T code, U status, V received, W mode, X notes
            add_dv(categories, "D")
            add_dv(weights, "F")
            add_dv(supplies, "G")
            add_dv(purities, "H")
            add_dv(sources, "I")
            add_dv(["pending", "delivered"], "U")
            add_dv(["cash", "upi", "bank_transfer", "old_gold_exchange", "other"], "W")
        elif title == "Cash Entries":
            add_dv(["received", "paid"], "D")
        elif title == "Opening Balances":
            add_dv(["customer", "party"], "A")
            add_dv(["debit", "credit"], "E")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
