"""Parse, validate, and commit a filled import template.

Validation runs first and reports every problem (with sheet + row) so the admin
can fix the file before anything is written. Commit applies all sheets in a single
transaction — any error rolls back the whole import. Customer/party names reuse the
same matching as manual entry (:mod:`app.services.matching`).
"""
from __future__ import annotations

import io
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import (
    CashEntry,
    DiamondType,
    ItemCategory,
    Order,
    OrderImage,
    OrderItem,
    OrderItemDiamond,
    OrderPayment,
    OrderSource,
    Purchase,
    PurityType,
    SupplySource,
    User,
    WeightType,
)
from app.models.base import (
    BalanceDirection,
    CashEntryType,
    EntityType,
    OrderStatus,
    PaymentMode,
)
from app.services.import_template import SHEETS
from app.services.seed import LEGACY_DIAMOND_TYPE
from app.services.ledger import set_opening_balance
from app.services.matching import (
    find_customer_match,
    find_party_match,
    get_or_create_customer,
    get_or_create_party,
)
from app.services.orders import compute_net_weight, compute_subtotal

# Numeric weight/rate columns on the Orders sheet (checked on every row).
_WEIGHT_FIELDS = ("gross_weight", "diamond_weight", "stone_weight", "others_weight")
_RATE_FIELDS = ("metal_rate", "diamond_rate", "stone_rate", "others_rate", "labour_rate")

VALID_STATUS = {s.value for s in OrderStatus}
VALID_PAYMENT = {m.value for m in PaymentMode}
VALID_CASH = {t.value for t in CashEntryType}
VALID_ENTITY = {e.value for e in EntityType}
VALID_DIRECTION = {d.value for d in BalanceDirection}

_IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")
_IMAGE_MIME = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
               ".webp": "image/webp", ".gif": "image/gif"}
MAX_IMAGES_PER_ITEM = 12


def split_upload(data: bytes) -> tuple[bytes, dict[str, bytes]]:
    """Split an upload into (workbook bytes, {filename_lower: bytes}).

    Accepts either a plain ``.xlsx`` or a ``.zip`` bundle containing the workbook
    plus an ``images/`` folder. (An ``.xlsx`` is itself a zip, so a bundle is
    detected by the presence of a ``.xlsx`` member inside it.)
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
        names = zf.namelist()
    except zipfile.BadZipFile:
        return data, {}
    xlsx_members = [n for n in names if n.lower().endswith(".xlsx")]
    if not xlsx_members:
        return data, {}  # a plain .xlsx workbook
    workbook = zf.read(xlsx_members[0])
    images: dict[str, bytes] = {}
    for n in names:
        if n.lower().endswith(_IMAGE_EXTS):
            images[_img_key(n)] = zf.read(n)
    return workbook, images


def _img_key(name: str) -> str:
    """Lowercased basename used to match a sheet filename to a bundled image."""
    return name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1].strip().lower()


def _pair_row(entry):
    """Row-dict accessor for grouping ``(row_number, row)`` pairs."""
    return entry[1]


def _group_consecutive(entries: list, key: str, get=lambda r: r) -> list[list]:
    """Group consecutive entries sharing a non-blank value for ``key``.

    Used twice on the Orders sheet: ``order_ref`` collapses rows into one
    multi-piece order, then ``item_ref`` collapses an order's rows into one piece
    carrying several typed diamond lines. A blank key always starts a new group,
    so unreferenced rows stay standalone (the pre-grouping behaviour).
    """
    groups: list[list] = []
    for entry in entries:
        val = _s(get(entry).get(key))
        if val and groups and _s(get(groups[-1][0]).get(key)) == val:
            groups[-1].append(entry)
        else:
            groups.append([entry])
    return groups


def _image_names(value) -> list[str]:
    """Parse the Orders 'images' cell into filenames (separated by ; or ,)."""
    return [p.strip() for p in _s(value).replace(",", ";").split(";") if p.strip()]


def _guess_mime(filename: str) -> str:
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    return _IMAGE_MIME.get(ext, "image/jpeg")


def _s(value) -> str:
    return "" if value is None else str(value).strip()


def _parse_date(value) -> Optional[date]:
    if value is None or _s(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_s(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_decimal(value) -> Optional[Decimal]:
    if value is None or _s(value) == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def parse_workbook(data: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    out: dict[str, list[dict]] = {}
    for title, headers in SHEETS.items():
        rows: list[dict] = []
        if title in wb.sheetnames:
            ws = wb[title]
            for raw in ws.iter_rows(min_row=2, values_only=True):
                if raw is None or all(c is None or _s(c) == "" for c in raw):
                    continue
                rows.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
        out[title] = rows
    return out


def validate(session: Session, sheets: dict[str, list[dict]], images: dict | None = None) -> dict:
    errors: list[dict] = []
    available = set(images or {})

    def err(sheet, row, msg):
        errors.append({"sheet": sheet, "row": row, "message": msg})

    purities = {p.name.strip().lower() for p in session.query(PurityType).all()}
    categories = {c.name.strip().lower() for c in session.query(ItemCategory).all()}
    weight_types = {w.name.strip().lower() for w in session.query(WeightType).all()}
    supply_sources = {s.name.strip().lower() for s in session.query(SupplySource).all()}
    order_sources = {s.name.strip().lower() for s in session.query(OrderSource).all()}
    diamond_types = {d.name.strip().lower() for d in session.query(DiamondType).all()}

    for label in ("Customers", "Parties"):
        for i, row in enumerate(sheets.get(label, []), start=2):
            if not _s(row.get("name")):
                err(label, i, "name is required")

    # Orders rows nest twice: order_ref groups rows into one multi-piece order, then
    # item_ref groups an order's rows into one piece with several typed diamond lines.
    # The head row of each group carries that level's fields; the rest are continuations.
    order_rows = list(enumerate(sheets.get("Orders", []), start=2))
    for ogroup in _group_consecutive(order_rows, "order_ref", _pair_row):
        i, head = ogroup[0]
        if not _s(head.get("customer_name")):
            err("Orders", i, "customer_name is required")
        if _parse_date(head.get("order_date")) is None:
            err("Orders", i, "order_date missing or not YYYY-MM-DD")
        src = _s(head.get("source")).lower()
        if src and src not in order_sources:
            err("Orders", i, f"unknown source '{head.get('source')}'")
        status = _s(head.get("status")).lower() or "pending"
        if status not in VALID_STATUS:
            err("Orders", i, f"invalid status '{status}'")
        mode = _s(head.get("payment_mode")).lower()
        if mode and mode not in VALID_PAYMENT:
            err("Orders", i, f"invalid payment_mode '{mode}'")
        if _parse_decimal(head.get("payment_received")) is None:
            err("Orders", i, "payment_received is not a number")

        for igroup in _group_consecutive(ogroup, "item_ref", _pair_row):
            pi, prow = igroup[0]
            category = _s(prow.get("item_category")).lower()
            if not category:
                err("Orders", pi, "item_category is required")
            elif category not in categories:
                err("Orders", pi, f"unknown item_category '{prow.get('item_category')}'")
            wt = _s(prow.get("weight_type")).lower()
            if wt and wt not in weight_types:
                err("Orders", pi, f"unknown weight_type '{prow.get('weight_type')}'")
            ss = _s(prow.get("supply_source")).lower()
            if ss and ss not in supply_sources:
                err("Orders", pi, f"unknown supply_source '{prow.get('supply_source')}'")
            purity = _s(prow.get("purity")).lower()
            if purity and purity not in purities:
                err("Orders", pi, f"unknown purity '{prow.get('purity')}'")
            # A piece's pictures may be listed on any of its rows; the cap is per piece.
            piece_images: list[str] = []
            for ri, row in igroup:
                names = _image_names(row.get("images"))
                for fname in names:
                    if not fname.lower().endswith(_IMAGE_EXTS):
                        err("Orders", ri, f"'{fname}' is not an image file (png/jpg/jpeg/webp/gif)")
                    elif _img_key(fname) not in available:
                        err("Orders", ri, f"image '{fname}' not found in the uploaded .zip")
                piece_images.extend(names)
            if len(piece_images) > MAX_IMAGES_PER_ITEM:
                err("Orders", pi, f"too many images ({len(piece_images)} > {MAX_IMAGES_PER_ITEM})")

        # Every row can carry its own diamond line + numeric weights/rates.
        for ri, row in ogroup:
            dt = _s(row.get("diamond_type")).lower()
            if dt and dt not in diamond_types:
                err("Orders", ri, f"unknown diamond_type '{row.get('diamond_type')}'")
            for fld in _WEIGHT_FIELDS + _RATE_FIELDS:
                if _parse_decimal(row.get(fld)) is None:
                    err("Orders", ri, f"{fld} is not a number")

    for i, row in enumerate(sheets.get("Cash Entries", []), start=2):
        if _parse_date(row.get("date")) is None:
            err("Cash Entries", i, "date missing or not YYYY-MM-DD")
        if _s(row.get("type")).lower() not in VALID_CASH:
            err("Cash Entries", i, "type must be received or paid")
        if _parse_decimal(row.get("amount")) is None:
            err("Cash Entries", i, "amount is not a number")

    for i, row in enumerate(sheets.get("Purchases", []), start=2):
        if not _s(row.get("party_name")):
            err("Purchases", i, "party_name is required")
        if _parse_date(row.get("date")) is None:
            err("Purchases", i, "date missing or not YYYY-MM-DD")
        for fld in ("amount", "amount_paid"):
            if _parse_decimal(row.get(fld)) is None:
                err("Purchases", i, f"{fld} is not a number")

    for i, row in enumerate(sheets.get("Opening Balances", []), start=2):
        if _s(row.get("entity_type")).lower() not in VALID_ENTITY:
            err("Opening Balances", i, "entity_type must be customer or party")
        if not _s(row.get("entity_name")):
            err("Opening Balances", i, "entity_name is required")
        if _s(row.get("direction")).lower() not in VALID_DIRECTION:
            err("Opening Balances", i, "direction must be debit or credit")
        if _parse_decimal(row.get("amount")) is None:
            err("Opening Balances", i, "amount is not a number")
        if _parse_date(row.get("as_of_date")) is None:
            err("Opening Balances", i, "as_of_date missing or not YYYY-MM-DD")

    summary = {title: len(rows) for title, rows in sheets.items()}
    if available:
        summary["Images"] = len(available)
    return {"ok": len(errors) == 0, "errors": errors, "summary": summary}


def commit(session: Session, user: User, sheets: dict[str, list[dict]],
           images: dict | None = None, today: Optional[date] = None) -> dict:
    """Apply a validated import in one transaction. Raises on error (rolled back)."""
    today = today or date.today()
    images = images or {}
    purity_by_name = {p.name.strip().lower(): p for p in session.query(PurityType).all()}
    category_by_name = {c.name.strip().lower(): c for c in session.query(ItemCategory).all()}
    weight_by_name = {w.name.strip().lower(): w for w in session.query(WeightType).all()}
    supply_by_name = {s.name.strip().lower(): s for s in session.query(SupplySource).all()}
    source_by_name = {s.name.strip().lower(): s for s in session.query(OrderSource).all()}
    diamond_by_name = {d.name.strip().lower(): d for d in session.query(DiamondType).all()}
    legacy_diamond = diamond_by_name.get(LEGACY_DIAMOND_TYPE.strip().lower())
    counts = {"customers": 0, "parties": 0, "orders": 0, "images": 0,
              "cash_entries": 0, "purchases": 0, "opening_balances": 0}

    try:
        for row in sheets.get("Customers", []):
            cust, created = get_or_create_customer(session, _s(row.get("name")), created_by=user.id)
            if created:
                cust.phone = _s(row.get("phone")) or None
                cust.address = _s(row.get("address")) or None
                cust.notes = _s(row.get("notes")) or None
                counts["customers"] += 1
        for row in sheets.get("Parties", []):
            party, created = get_or_create_party(session, _s(row.get("name")), created_by=user.id)
            if created:
                party.phone = _s(row.get("phone")) or None
                party.address = _s(row.get("address")) or None
                party.notes = _s(row.get("notes")) or None
                counts["parties"] += 1

        # Orders rows nest twice: order_ref groups rows into one multi-piece order,
        # then item_ref groups an order's rows into one piece carrying several typed
        # diamond lines. Each group's head row holds that level's fields. A single
        # payment line is recorded from the order head row's amount + mode.
        for group in _group_consecutive(sheets.get("Orders", []), "order_ref"):
            head = group[0]
            customer, _ = get_or_create_customer(session, _s(head.get("customer_name")), created_by=user.id)
            order_date = _parse_date(head.get("order_date"))
            src = _s(head.get("source")).lower()
            mode = _s(head.get("payment_mode")).lower()
            received = _parse_decimal(head.get("payment_received")) or Decimal("0")
            order = Order(
                customer_id=customer.id,
                order_date=order_date,
                order_code=_s(head.get("order_code")) or None,
                notes=_s(head.get("notes")) or None,
                reference=_s(head.get("reference")) or None,
                source_id=source_by_name[src].id if src else None,
                status=OrderStatus(_s(head.get("status")).lower() or "pending"),
                payment_received=received,
                payment_mode=PaymentMode(mode) if mode else None,
                created_by=user.id,
                is_backdated=order_date < today if order_date else False,
                total_amount=Decimal("0"),
                balance=Decimal("0"),
            )
            session.add(order)
            session.flush()
            total = Decimal("0")
            for sort_order, igroup in enumerate(_group_consecutive(group, "item_ref")):
                prow = igroup[0]
                wt = _s(prow.get("weight_type")).lower()
                ss = _s(prow.get("supply_source")).lower()
                pu = _s(prow.get("purity")).lower()
                piece = OrderItem(
                    order_id=order.id,
                    item_category_id=category_by_name[_s(prow.get("item_category")).lower()].id,
                    item_name=_s(prow.get("item_name")) or None,
                    weight_type_id=weight_by_name[wt].id if wt else None,
                    supply_source_id=supply_by_name[ss].id if ss else None,
                    purity_type_id=purity_by_name[pu].id if pu else None,
                    gross_weight=_parse_decimal(prow.get("gross_weight")) or None,
                    stone_weight=_parse_decimal(prow.get("stone_weight")) or None,
                    others_weight=_parse_decimal(prow.get("others_weight")) or None,
                    metal_rate=_parse_decimal(prow.get("metal_rate")) or None,
                    stone_rate=_parse_decimal(prow.get("stone_rate")) or None,
                    others_rate=_parse_decimal(prow.get("others_rate")) or None,
                    labour_rate=_parse_decimal(prow.get("labour_rate")) or None,
                    sort_order=sort_order,
                )
                # Each row of the piece adds one typed diamond line; a blank
                # diamond_type falls back to the legacy "Other fancy" bucket.
                for row in igroup:
                    dia_ct = _parse_decimal(row.get("diamond_weight")) or Decimal("0")
                    dia_rate = _parse_decimal(row.get("diamond_rate")) or Decimal("0")
                    if dia_ct <= 0 and dia_rate <= 0:
                        continue
                    dt = _s(row.get("diamond_type")).lower()
                    dtype = diamond_by_name[dt] if dt else legacy_diamond
                    piece.diamonds.append(OrderItemDiamond(
                        diamond_type_id=dtype.id if dtype else None,
                        carats=dia_ct, rate=dia_rate or None,
                        sort_order=len(piece.diamonds),
                    ))
                net = compute_net_weight(piece)
                piece.net_weight = net
                piece.subtotal = compute_subtotal(piece, net)
                session.add(piece)
                session.flush()  # need piece.id to attach its pictures
                # A piece's pictures may be listed on any of its rows.
                names = [n for row in igroup for n in _image_names(row.get("images"))]
                for idx, fname in enumerate(names):
                    blob = images.get(_img_key(fname))
                    if blob is None:
                        continue
                    session.add(OrderImage(
                        order_item_id=piece.id,
                        filename=fname.rsplit("/", 1)[-1].rsplit("\\", 1)[-1],
                        mime=_guess_mime(fname),
                        data=blob,
                        sort_order=idx,
                    ))
                    counts["images"] += 1
                total += piece.subtotal
            if received > 0:
                session.add(OrderPayment(
                    order_id=order.id,
                    mode=PaymentMode(mode) if mode else PaymentMode.cash,
                    amount=received,
                    sort_order=0,
                ))
            order.total_amount = total
            order.balance = total - received
            counts["orders"] += 1

        for row in sheets.get("Cash Entries", []):
            person = _s(row.get("person_name"))
            cust = find_customer_match(session, person) if person else None
            party = find_party_match(session, person) if person else None
            cash_date = _parse_date(row.get("date"))
            session.add(CashEntry(
                entry_date=cash_date,
                person_name=person,
                customer_id=cust.id if cust else None,
                party_id=party.id if (party and not cust) else None,
                details=_s(row.get("details")) or None,
                entry_type=CashEntryType(_s(row.get("type")).lower()),
                amount=_parse_decimal(row.get("amount")) or Decimal("0"),
                created_by=user.id,
                is_backdated=cash_date < today if cash_date else False,
            ))
            counts["cash_entries"] += 1

        for row in sheets.get("Purchases", []):
            party, _ = get_or_create_party(session, _s(row.get("party_name")), created_by=user.id)
            amount = _parse_decimal(row.get("amount")) or Decimal("0")
            paid = _parse_decimal(row.get("amount_paid")) or Decimal("0")
            p_date = _parse_date(row.get("date"))
            session.add(Purchase(
                purchase_date=p_date,
                party_id=party.id,
                details=_s(row.get("details")) or None,
                entry_notes=_s(row.get("entry_notes")) or None,
                amount=amount,
                amount_paid=paid,
                balance=amount - paid,
                created_by=user.id,
                is_backdated=p_date < today if p_date else False,
            ))
            counts["purchases"] += 1

        for row in sheets.get("Opening Balances", []):
            etype = EntityType(_s(row.get("entity_type")).lower())
            name = _s(row.get("entity_name"))
            if etype == EntityType.customer:
                entity, _ = get_or_create_customer(session, name, created_by=user.id)
            else:
                entity, _ = get_or_create_party(session, name, created_by=user.id)
            set_opening_balance(
                session, etype, entity.id, _parse_date(row.get("as_of_date")),
                _parse_decimal(row.get("amount")) or Decimal("0"),
                BalanceDirection(_s(row.get("direction")).lower()), created_by=user.id,
            )
            counts["opening_balances"] += 1

        session.commit()
        return {"ok": True, "imported": counts}
    except Exception:
        session.rollback()
        raise
