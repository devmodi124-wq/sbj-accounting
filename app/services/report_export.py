"""Excel (.xlsx) export for reports — all fields, with optional product thumbnails.

Mirrors the CSV columns but, when a row carries a picture (Sales / Stock), embeds
a small thumbnail in a trailing "Picture" column. Requires Pillow to read/resize
images; rows without a usable picture simply leave the cell blank.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill

try:  # Pillow is needed to decode + resize the thumbnails.
    from PIL import Image as PILImage
except Exception:  # pragma: no cover - Pillow is a declared dependency
    PILImage = None

HEADER_FILL = PatternFill("solid", fgColor="1C1B19")
HEADER_FONT = Font(color="FBF8F2", bold=True)
THUMB_BOX = (72, 72)


def _thumbnail(data: bytes):
    """Return (PNG BytesIO, (w, h)) for a resized thumbnail, or None on failure."""
    if PILImage is None or not data:
        return None
    try:
        im = PILImage.open(BytesIO(data))
        im.thumbnail(THUMB_BOX)
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        size = im.size
        out = BytesIO()
        im.save(out, format="PNG")
        out.seek(0)
        return out, size
    except Exception:
        return None


def build_xlsx(
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict],
    thumbnails: dict | None = None,
) -> bytes:
    """Serialize ``rows`` to an .xlsx. ``columns`` is (key, header); ``thumbnails``
    maps a row's ``id`` to raw image bytes (adds a trailing Picture column)."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]

    headers = [h for _, h in columns]
    has_pics = thumbnails is not None
    if has_pics:
        headers = headers + ["Picture"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) + 2)

    pic_col = len(columns) + 1
    pic_letter = ws.cell(row=1, column=pic_col).column_letter if has_pics else None

    for r, row in enumerate(rows, start=2):
        for col, (key, _) in enumerate(columns, start=1):
            ws.cell(row=r, column=col, value=row.get(key, ""))
        if has_pics:
            thumb = _thumbnail((thumbnails or {}).get(row.get("id")))
            if thumb is not None:
                buf, (w, h) = thumb
                try:
                    ws.add_image(XLImage(buf), f"{pic_letter}{r}")
                    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, h * 0.75)
                    ws.column_dimensions[pic_letter].width = max(
                        ws.column_dimensions[pic_letter].width or 14, w / 7
                    )
                except Exception:
                    pass

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
