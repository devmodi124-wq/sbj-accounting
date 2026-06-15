"""Report filters + Excel (.xlsx) export with product thumbnails."""
from __future__ import annotations

import io
import zipfile
from datetime import date

from openpyxl import load_workbook

TODAY = date.today().isoformat()
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000a49444154789c6360000002000154a24f600000000049454e44ae426082"
)


def _order(client, **over):
    cat = client.get("/api/item-categories").json()[0]["id"]
    item = {"item_category_id": cat, "item_name": "Ring", "gross_weight": "1000", "metal_rate": "1"}
    payload = {"customer_name": "Rep Cust", "order_date": TODAY, "status": "delivered",
               "payments": [], "items": [item]}
    payload.update(over)
    return client.post("/api/orders", json=payload).json()


def test_sales_filter_by_source(admin_client):
    sources = admin_client.get("/api/order-sources").json()
    s0, s1 = sources[0]["id"], sources[1]["id"]
    _order(admin_client, source_id=s0)
    _order(admin_client, source_id=s1)
    filtered = admin_client.get("/api/reports/sales", params={"source_id": s0}).json()
    assert filtered["total"] == 1


def test_sales_filter_by_weight_type(admin_client):
    cat = admin_client.get("/api/item-categories").json()[0]["id"]
    weights = admin_client.get("/api/weight-types").json()
    w0, w1 = weights[0]["id"], weights[1]["id"]
    _order(admin_client, items=[{"item_category_id": cat, "weight_type_id": w0,
                                 "gross_weight": "1", "metal_rate": "1"}])
    _order(admin_client, items=[{"item_category_id": cat, "weight_type_id": w1,
                                 "gross_weight": "1", "metal_rate": "1"}])
    filtered = admin_client.get("/api/reports/sales", params={"weight_type_id": w0}).json()
    assert filtered["total"] == 1


def test_stock_filter_by_category(admin_client):
    cats = admin_client.get("/api/item-categories").json()
    ring, necklace = cats[0]["id"], cats[1]["id"]
    _order(admin_client, items=[{"item_category_id": ring, "gross_weight": "1", "metal_rate": "1"}])
    _order(admin_client, items=[{"item_category_id": necklace, "gross_weight": "1", "metal_rate": "1"}])
    filtered = admin_client.get("/api/reports/stock", params={"category_id": ring}).json()
    assert filtered["total"] == 1
    assert filtered["rows"][0]["item_category"] == cats[0]["name"]


def test_sales_xlsx_export_has_headers(admin_client):
    _order(admin_client)
    r = admin_client.get("/api/reports/sales", params={"format": "xlsx"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX_MIME)
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Customer" in headers and "Source" in headers and "Picture" in headers


def _real_png() -> bytes:
    from PIL import Image
    buf = io.BytesIO()
    Image.new("RGB", (12, 12), (168, 113, 74)).save(buf, format="PNG")
    return buf.getvalue()


def test_sales_xlsx_embeds_product_picture(admin_client):
    created = _order(admin_client)
    oid, iid = created["id"], created["items"][0]["id"]
    admin_client.post(f"/api/orders/{oid}/items/{iid}/images",
                      files=[("files", ("p.png", _real_png(), "image/png"))])
    r = admin_client.get("/api/reports/sales", params={"format": "xlsx"})
    # openpyxl doesn't round-trip images on read; check the xlsx package instead.
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    assert any(n.startswith("xl/media/") for n in zf.namelist())   # thumbnail embedded


def test_other_reports_xlsx(admin_client):
    _order(admin_client, payments=[{"mode": "upi", "amount": "100"}])  # leaves a balance
    for name in ("debtors", "customers", "stock"):
        r = admin_client.get(f"/api/reports/{name}", params={"format": "xlsx"})
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(XLSX_MIME)
        load_workbook(io.BytesIO(r.content))  # parses without error
