"""Phase 8 — Excel import: template, validation, commit (+ picture ZIP bundles)."""
from __future__ import annotations

import io
import zipfile

from openpyxl import Workbook, load_workbook

from app.services.import_template import SHEETS

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000a49444154789c6360000002000154a24f600000000049454e44ae426082"
)


def make_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers in SHEETS.items():
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in data.get(title, []):
            ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_zip(data: dict, images: dict[str, bytes]) -> bytes:
    """Bundle a filled workbook + an images/ folder into a .zip upload."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("khata-import.xlsx", make_xlsx(data))
        for name, blob in images.items():
            zf.writestr(f"images/{name}", blob)
    return buf.getvalue()


def _upload(client, url, content):
    return client.post(url, files={"file": ("import.xlsx", content, XLSX)})


GOOD = {
    "Customers": [{"name": "Imp Cust", "phone": "123"}],
    "Orders": [{"order_ref": "O1", "customer_name": "Imp Cust", "order_date": "2026-06-01",
                "item_category": "Ring", "item_name": "Ladies ring", "weight_type": "Normal",
                "supply_source": "On Order", "purity": "22 KT", "gross_weight": "1200",
                "metal_rate": "1", "status": "delivered", "payment_received": "500",
                "payment_mode": "cash"}],
    "Cash Entries": [{"date": "2026-06-02", "person_name": "Imp Cust", "type": "received", "amount": "300"}],
    "Purchases": [{"date": "2026-06-03", "party_name": "Imp Supp", "amount": "5000", "amount_paid": "1000"}],
    "Opening Balances": [{"entity_type": "customer", "entity_name": "Imp Cust",
                          "as_of_date": "2026-01-01", "amount": "100", "direction": "debit"}],
}


def test_template_download(admin_client):
    r = admin_client.get("/api/import/template")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "Instructions" in wb.sheetnames
    assert "Orders" in wb.sheetnames


def test_validate_good_file(admin_client):
    r = _upload(admin_client, "/api/import/validate", make_xlsx(GOOD))
    body = r.json()
    assert body["ok"] is True
    assert body["errors"] == []
    assert body["summary"]["Orders"] == 1


def test_validate_reports_errors(admin_client):
    bad = {
        "Orders": [{"order_ref": "O1", "customer_name": "", "order_date": "nope",
                    "item_name": "", "status": "weird", "payment_received": "x",
                    "gross_weight": "notnum"}],
    }
    body = _upload(admin_client, "/api/import/validate", make_xlsx(bad)).json()
    assert body["ok"] is False
    msgs = " ".join(e["message"] for e in body["errors"])
    assert "customer_name is required" in msgs
    assert "item_category is required" in msgs
    assert "order_date" in msgs
    assert "payment_received is not a number" in msgs
    assert "gross_weight is not a number" in msgs


def test_commit_imports_everything(admin_client):
    r = _upload(admin_client, "/api/import/commit", make_xlsx(GOOD))
    assert r.status_code == 200
    imported = r.json()["imported"]
    assert imported["orders"] == 1
    assert imported["purchases"] == 1

    # Order priced from gross 1200 × metal_rate 1 = 1200; balance 1200 - 500.
    orders = admin_client.get("/api/orders").json()
    assert orders[0]["total_amount"] == "1200.00"
    assert orders[0]["balance"] == "700.00"

    # Customer ledger shows the opening balance and the order.
    cid = orders[0]["customer_id"]
    led = admin_client.get(f"/api/ledgers/customer/{cid}").json()
    assert led["entries"][0]["particulars"] == "Opening balance"


def test_commit_groups_multi_item_order(admin_client):
    # Two rows sharing order_ref MULTI collapse into one order with two pieces;
    # order-level fields live on the first row only, totals sum both subtotals.
    data = {
        "Orders": [
            {"order_ref": "MULTI", "customer_name": "Multi Cust", "order_date": "2026-06-01",
             "item_category": "Ring", "item_name": "Ring A", "gross_weight": "100",
             "metal_rate": "1", "status": "delivered", "payment_received": "120",
             "payment_mode": "cash"},
            {"order_ref": "MULTI", "item_category": "Ring", "item_name": "Ring B",
             "gross_weight": "50", "metal_rate": "1"},
        ],
    }
    r = _upload(admin_client, "/api/import/commit", make_xlsx(data))
    assert r.status_code == 200
    assert r.json()["imported"]["orders"] == 1

    orders = admin_client.get("/api/orders").json()
    assert len(orders) == 1
    # 100×1 + 50×1 = 150 total; balance 150 − 120.
    assert orders[0]["total_amount"] == "150.00"
    assert orders[0]["balance"] == "30.00"

    detail = admin_client.get(f"/api/orders/{orders[0]['id']}").json()
    names = [it["item_name"] for it in detail["items"]]
    assert names == ["Ring A", "Ring B"]


def test_commit_rejected_when_invalid(admin_client):
    bad = {"Orders": [{"order_ref": "O1", "customer_name": "", "order_date": "x", "item_name": ""}]}
    r = _upload(admin_client, "/api/import/commit", make_xlsx(bad))
    assert r.status_code == 422


def test_commit_reuses_existing_customer(admin_client):
    admin_client.post("/api/customers", json={"name": "Existing Cust"})
    data = {
        "Orders": [{"order_ref": "X1", "customer_name": "  existing cust ", "order_date": "2026-06-01",
                    "item_category": "Ring", "status": "pending", "gross_weight": "100",
                    "metal_rate": "1", "payment_received": "0"}],
    }
    _upload(admin_client, "/api/import/commit", make_xlsx(data))
    matches = admin_client.get("/api/customers", params={"q": "existing cust"}).json()
    assert len([c for c in matches if c["name"].lower() == "existing cust"]) == 1


def _zip_upload(client, url, content):
    return client.post(url, files={"file": ("import.zip", content, "application/zip")})


def test_zip_bundle_imports_pictures(admin_client):
    data = {"Orders": [{"order_ref": "P1", "customer_name": "Pic Cust", "order_date": "2026-06-01",
                        "item_category": "Ring", "gross_weight": "100", "metal_rate": "1",
                        "status": "delivered", "images": "ring1.png;ring2.png"}]}
    content = make_zip(data, {"ring1.png": PNG, "ring2.png": PNG})

    assert _zip_upload(admin_client, "/api/import/validate", content).json()["ok"] is True
    imported = _zip_upload(admin_client, "/api/import/commit", content).json()["imported"]
    assert imported["orders"] == 1
    assert imported["images"] == 2

    order = admin_client.get("/api/orders").json()[0]
    oid = order["id"]
    iid = admin_client.get(f"/api/orders/{oid}").json()["items"][0]["id"]
    imgs = admin_client.get(f"/api/orders/{oid}/items/{iid}/images").json()
    assert len(imgs) == 2


def test_missing_picture_reported(admin_client):
    data = {"Orders": [{"order_ref": "P1", "customer_name": "Pic Cust", "order_date": "2026-06-01",
                        "item_category": "Ring", "gross_weight": "100", "metal_rate": "1",
                        "images": "present.png;missing.png"}]}
    content = make_zip(data, {"present.png": PNG})   # missing.png not bundled
    body = _zip_upload(admin_client, "/api/import/validate", content).json()
    assert body["ok"] is False
    assert any("missing.png" in e["message"] for e in body["errors"])


def test_plain_xlsx_with_image_names_but_no_zip_errors(admin_client):
    # Listing images in a plain .xlsx (no bundle) should fail validation, not crash.
    data = {"Orders": [{"order_ref": "P1", "customer_name": "Pic Cust", "order_date": "2026-06-01",
                        "item_category": "Ring", "gross_weight": "100", "metal_rate": "1",
                        "images": "ring1.png"}]}
    body = _upload(admin_client, "/api/import/validate", make_xlsx(data)).json()
    assert body["ok"] is False
    assert any("not found" in e["message"] for e in body["errors"])


def test_import_requires_admin(admin_client):
    from fastapi.testclient import TestClient

    from app.main import app

    admin_client.post("/api/users", json={"username": "emp", "password": "emp123"})
    emp = TestClient(app)
    emp.post("/auth/login", json={"username": "emp", "password": "emp123"})
    assert emp.get("/api/import/template").status_code == 403
